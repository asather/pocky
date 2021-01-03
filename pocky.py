import os.path
from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import json
import requests

class pocky:

    def __init__(self, keys):

        self.settings = self.load_settings(keys)
        self.base_send = 'https://getpocket.com/v3/send'
        self.base_retr = 'https://getpocket.com/v3/get'
        self.base_add = 'https://getpocket.com/v3/add'
        self.debug = True

    def debugger(self, message):
        #Prints debugger message.

        if(self.debug == False):
            return False

        final_message = ""
        if(type(message) is list):
            for i in message:
                final_message = final_message + ' ' + str(i)
        else:
            final_message = str(message)
        print('DEBUGGER: ' + final_message)


    def get_sources(self):
        #Returns JSON of the news sources.

        return self.settings['sources']


    def package_action(self, action, item_id):
        #Packages the action in the specific string for inclusion in the URL.
        #Basically provides encoding for a list of python dictionaries.

        #%5B%7B%22action%22%3A%22archive%22%2C%22item_id%22%3A229279689%7D%5D

        action = "actions=%5B%7B%22action%22%3A%22" + action + "%22%2C%22item_id%22%3A" + str(item_id) + "%7D%5D"
        url = self.base_send + '?' + action + '&' + self.get_keys()
        return url

    def get_keys(self):
        #Returns the consumer key and access token, formatted for a URL.

        return 'consumer_key=' + self.settings['consumer_key'] + '&access_token=' + self.settings['access_token']

    def execute(self, url, parameters = False):
        #Makes the request to Pocket's API endpoints.

        self.debugger(url)

        #Handle submission.
        if(parameters != False):
            self.debugger(parameters)
            r = requests.post(url, params = parameters)
        else:
            r = requests.post(url)

        #Handle response.
        if(r.status_code == 200):
            self.debugger(r.json())
            return r
        else:
            self.debugger('FAILURE')
            print(r.text)
            print(r.url)
            return False

    def archive(self, item_id):
        #Archive an article ID.

        url = self.package_action('archive', item_id)
        return self.execute(url)


    def retrieve(self, state, count = False):
        #Returns all articles in a given state.
        #Read more: https://getpocket.com/developer/docs/v3/retrieve

        allowed_states = ['unread', 'archive', 'all']
        if(state.lower() not in allowed_states):
            return {'result' : False, 'error' : 'State "' + state + '" is not allowed.'}

        parameters = {
            'state' : state.lower(),
            'consumer_key':self.settings['consumer_key'],
            'access_token':self.settings['access_token']
        }

        #Add Count to the object if it's provided.
        if(count != False):
            parameters['count'] = count

        #Execute...
        r = self.execute(self.base_retr, parameters)

        #Return JSON so long as it's needed...
        if(r != False):
            return r.json()
        else:
            return r

    def load_history(self):
        #Opens the history of items added.

        unc = Path(self.settings['history_file'])
        f = open(unc)
        content = f.read()
        f.close()

        return content.splitlines()

    def add_history(self, item):
        #Adds a URL to the end.

        unc = Path(self.settings['history_file'])
        f = open(unc, "a")
        f.write("\n" + item)
        f.close()

    def add(self, url, tags = False):
        #Add an item to Pocket.
        #Tags is optional and must be a list.

        parameters = {
            'url' : url,
            'consumer_key':self.settings['consumer_key'],
            'access_token':self.settings['access_token']
        }

        #Add tags to the object if it's provided.
        if(tags != False):
            parameters['tags'] = tags

        self.execute(self.base_add, parameters)


    def load_settings(self, filepath):
        #Opens the config json file and loads settings into a dictionary.

        settings_dict = {}
        with open(filepath) as json_file:
        	data = json.load(json_file)
        return data

    def data_from_xlsx(self, unc, sheet_name):
        #Returns a Python list of values for a given worksheet within a workbook..
        #Must have a header row.
        #If this is to be expanded in the future to allow for ranges without header rows,
        #look into cell.letter property of the python Excel object.

        #Open the workbook
        workbook = load_workbook(unc)

        #Check if sheet exists.
        if(sheet_name not in workbook.sheetnames):
            self.debugger('The sheet ' + sheet_name + ' is not in the workbook at ' + unc)
            return False

        #Open the worksheet.
        worksheet = workbook[sheet_name]

        #Declare the returned data object.
        worksheet_data = []

        #Get the header.
        header_row_values = [cell.value for cell in worksheet[1]]
        self.debugger('Sheet open. Header values are as follows: ' + str(header_row_values))

        #Loop through all rows but the first one.
        first_row = True
        for row in worksheet.rows:

            if(first_row): #ignore the header row.
                first_row = False
                continue

            values = {}
            for key, cell in zip(header_row_values, row):
                values[key] = cell.value
            worksheet_data.append(values)

        return worksheet_data

    def data_to_xlsx(self, data, header, unc, sheet_name):
        #Accepts data and the workbook object to add the data to
        #Prints the data into Exel.
        #Data must be a list of dictionaries.
        #Header is a list of column headers. They should be keys
        #in the dictionaries contained in data.
        #sheet_name should be the name of the sheet you wish to add the data to.
        #If the sheet does not exist, it will be created.
        #If a workbook exists at the UNC, it will be updated. If the workbook
        #does not exist at the UNC, it will be created.

        #If the workbook does not exist, create it.
        #otherwise, load it.
        if(os.path.exists(unc) == True):
            workbook = load_workbook(unc)
        else:
            workbook = Workbook()
            workbook.save(unc)


        #Check if sheet_name exists.
        if(sheet_name not in workbook.sheetnames):
            workbook.create_sheet(sheet_name)

        #Set the worksheet.
        worksheet = workbook[sheet_name]

        #Set starting markers.
        row = 1
        col = 1

        #Create header.
        for h in header:
            #print(h)
            worksheet.cell(row, col, h)
            col = col + 1

        row = row + 1 #Next row, after the header.

        #Write the data.
        for d in data: #Loop through each dictionary.
            col = 1 #Reset col for every row.
            for h in header:
                #Skip if this dictionary doesn't have that header.
                if(h in d.keys()):
                    #print(d[h])
                    worksheet.cell(row, col, d[h])
                else:
                    self.debugger('Skipping column ' + h + ' as it does not exist for this data dictionary.')
                    #print(d)
                col = col + 1
            row = row + 1

        workbook.save(unc)
        return workbook
